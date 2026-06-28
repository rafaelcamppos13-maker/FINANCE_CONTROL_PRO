from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# ===========================
# CORES
# ===========================

AZUL_PRIMARIO = "0F4C81"
AZUL_CLARO = "1E88E5"
BRANCO = "FFFFFF"
CINZA = "EAEAEA"
VERDE = "4CAF50"
VERMELHO = "F44336"
AMARELO = "FFC107"

# ===========================
# PREENCHIMENTOS
# ===========================

FILL_AZUL = PatternFill(
    fill_type="solid",
    fgColor=AZUL_PRIMARIO
)

FILL_CINZA = PatternFill(
    fill_type="solid",
    fgColor=CINZA
)

# ===========================
# FONTES
# ===========================

FONTE_TITULO = Font(
    name="Calibri",
    size=20,
    bold=True,
    color=BRANCO
)

FONTE_SUB = Font(
    name="Calibri",
    size=12,
    bold=True,
    color=BRANCO
)

FONTE_NORMAL = Font(
    name="Calibri",
    size=11
)

# ===========================
# BORDAS
# ===========================

BORDA = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# ===========================
# ALINHAMENTO
# ===========================

CENTRO = Alignment(
    horizontal="center",
    vertical="center"
)