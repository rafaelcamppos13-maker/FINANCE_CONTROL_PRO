from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path(__file__).parent
EXCEL = BASE / "excel"

EXCEL.mkdir(exist_ok=True)

arquivo = EXCEL / "FINANCE_CONTROL_PRO.xlsx"

wb = Workbook()
ws = wb.active

ws.title = "Dashboard"

# largura
ws.column_dimensions["A"].width = 40

# titulo
ws.merge_cells("A1:F3")
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 30
ws.row_dimensions[3].height = 15

titulo = ws["A1"]

titulo.value = "FINANCE CONTROL PRO"

titulo.font = Font(
    size=18,
    bold=True,
    color="FFFFFF"
)

titulo.fill = PatternFill(
    fill_type="solid",
    start_color="0F172A"
)

titulo.alignment = Alignment(
    horizontal="center",
    vertical="center"
)

# mensagem
ws["A4"] = "Sistema inicial criado com sucesso"

wb.save(arquivo)

print("Dashboard estilizado criado")